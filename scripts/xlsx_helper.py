import json
import sys
from pathlib import Path
from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import PatternFill, Font

DESC_KEYS = {'nghiep vu', 'nghiệp vụ', 'dien giai', 'diễn giải', 'mo ta', 'mô tả', 'description', 'transaction'}
DATE_KEYS = {'ngay', 'ngày', 'posting date', 'date'}
DOC_KEYS = {'so ct', 'số ct', 'so chung tu', 'số chứng từ', 'so_ct', 'voucher', 'reference', 'ref'}
AMOUNT_KEYS = {'so tien', 'số tiền', 'amount', 'value'}
QTY_KEYS = {'so luong', 'số lượng', 'sl', 'qty', 'quantity'}
UNIT_PRICE_KEYS = {'don gia', 'đơn giá', 'unit price', 'price each', 'gia mua', 'gia ban'}
OBJECT_KEYS = {'doi tuong', 'đối tượng', 'customer', 'vendor', 'object', 'counterparty', 'khach hang', 'nha cung cap'}
ITEM_KEYS = {'mat hang', 'mặt hàng', 'item', 'sku', 'product', 'hang hoa'}


def norm(v):
    return ' '.join(str(v or '').strip().lower().replace('_', ' ').replace('-', ' ').split())


def read_cmd(path_str: str):
    path = Path(path_str)
    wb = load_workbook(path, data_only=True)
    target = None
    for ws in wb.worksheets:
        if any(k in norm(ws.title) for k in ['input', 'nghiep vu', 'nghiệp vụ', 'transaction', 'journal', 'nhat ky']):
            target = ws
            break
    if target is None:
        target = wb.worksheets[0]

    rows = list(target.iter_rows(values_only=True))
    nonempty = [list(r) for r in rows if any(c is not None and str(c).strip() != '' for c in r)]
    if not nonempty:
        print(json.dumps({'transactions': [], 'records': [], 'headers': [], 'rawRows': []}, ensure_ascii=False))
        return

    header = [norm(c) for c in nonempty[0]]

    def find_idx(keys):
        for i, h in enumerate(header):
            if h in keys or any(k in h for k in keys):
                return i
        return -1

    desc_idx = find_idx(DESC_KEYS)
    date_idx = find_idx(DATE_KEYS)
    doc_idx = find_idx(DOC_KEYS)
    amount_idx = find_idx(AMOUNT_KEYS)
    qty_idx = find_idx(QTY_KEYS)
    unit_price_idx = find_idx(UNIT_PRICE_KEYS)
    object_idx = find_idx(OBJECT_KEYS)
    item_idx = find_idx(ITEM_KEYS)

    raw_rows = []
    for i, row in enumerate(nonempty[1:], start=2):
        vals = [str(c).strip() if c is not None else '' for c in row]
        while len(vals) < len(header):
            vals.append('')
        raw_rows.append({'rowNumber': i, 'cells': vals, 'map': {header[j]: vals[j] for j in range(len(header))}})

    records = []
    txns = []
    for rr in raw_rows:
        vals = rr['cells']
        if desc_idx >= 0 and desc_idx < len(vals):
            description = vals[desc_idx].strip()
        else:
            description = ' '.join(v for v in vals if v).strip()
        if not description:
            continue
        record = {
            'rowNumber': rr['rowNumber'],
            'date': vals[date_idx].strip() if 0 <= date_idx < len(vals) else '',
            'documentNo': vals[doc_idx].strip() if 0 <= doc_idx < len(vals) else '',
            'description': description,
            'amountHint': vals[amount_idx].strip() if 0 <= amount_idx < len(vals) else '',
            'quantity': vals[qty_idx].strip() if 0 <= qty_idx < len(vals) else '',
            'unitPrice': vals[unit_price_idx].strip() if 0 <= unit_price_idx < len(vals) else '',
            'object': vals[object_idx].strip() if 0 <= object_idx < len(vals) else '',
            'item': vals[item_idx].strip() if 0 <= item_idx < len(vals) else ''
        }
        records.append(record)
        txns.append(description)

    print(json.dumps({'transactions': txns, 'records': records, 'sheet': target.title, 'headers': header, 'rawRows': raw_rows}, ensure_ascii=False))


def write_cmd(payload_path: str):
    data = json.loads(Path(payload_path).read_text(encoding='utf-8'))
    out = Path(data['output'])
    out.parent.mkdir(parents=True, exist_ok=True)
    wb = Workbook()
    default = wb.active
    wb.remove(default)
    for sheet in data['sheets']:
      ws = wb.create_sheet(title=str(sheet['name'])[:31] or 'Sheet1')
      for row in sheet['rows']:
          ws.append(list(row))
      if sheet['rows']:
          ws.freeze_panes = 'A2'
          header_fill = PatternFill(fill_type='solid', fgColor='111111')
          soft_red = PatternFill(fill_type='solid', fgColor='FDECEC')
          soft_yellow = PatternFill(fill_type='solid', fgColor='FFF7D6')
          soft_green = PatternFill(fill_type='solid', fgColor='E9F8EE')
          for cell in ws[1]:
              cell.fill = header_fill
              cell.font = Font(color='FFFFFF', bold=True)
          # highlight status/level columns
          header_vals = [str(c.value or '').strip().lower() for c in ws[1]]
          for row in ws.iter_rows(min_row=2):
              for cell, header in zip(row, header_vals):
                  val = str(cell.value or '').strip().lower()
                  if header in ('status','mức độ','muc do','level'):
                      if val in ('review','high'):
                          cell.fill = soft_red
                      elif val in ('fixed','medium'):
                          cell.fill = soft_yellow
                      elif val in ('ok','low'):
                          cell.fill = soft_green
          for idx, col in enumerate(ws.columns, start=1):
              max_len = 0
              for cell in col:
                  value = '' if cell.value is None else str(cell.value)
                  if len(value) > max_len:
                      max_len = len(value)
              ws.column_dimensions[get_column_letter(idx)].width = min(max(max_len + 2, 12), 40)
    wb.save(out)
    print(json.dumps({'ok': True, 'output': str(out)}, ensure_ascii=False))


if __name__ == '__main__':
    cmd = sys.argv[1]
    if cmd == 'read':
        read_cmd(sys.argv[2])
    elif cmd == 'write':
        write_cmd(sys.argv[2])
    else:
        raise SystemExit('unknown command')
